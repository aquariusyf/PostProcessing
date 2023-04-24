from PostProcessingUtils import PostProcessingUtils
from openpyxl import Workbook, load_workbook
from datetime import datetime
from FilterMask import *
import sys
import os

filter_mask[ANALYZER_FILTER] = [';Common Displays;VoIP Analysis;VONR MO Call Setup Summary',
                                ';Common Displays;VoIP Analysis;VONR MT Call Setup Summary']

ExtractCSD = PostProcessingUtils()
ExtractCSD.getArgv(sys.argv)
ExtractCSD.scanWorkingDir()
ExtractCSD.exportAnalyzer('_CSD_All_Grid.xlsm')
ExtractCSD.scanWorkingDir('.xlsm')
ALL_CSD_Files = ExtractCSD.getFilesPath()

MO_First_Row_Call1 = ['MO Call Setup Delay', 'AVG']
MO_First_Row_Call2 = ['MO Call Setup Delay', 'AVG']
MO_Summary_Sheet_Call1 = {}
MO_Summary_Sheet_Call2 = {}
MT_First_Row_Call1 = ['MT Call Setup Delay', 'AVG']
MT_First_Row_Call2 = ['MT Call Setup Delay', 'AVG']
MT_Summary_Sheet_Call1 = {}
MT_Summary_Sheet_Call2 = {}
isFirstFile = True

for CSD_file in ALL_CSD_Files:
    logName = os.path.split(CSD_file)[1]
    if not logName.endswith('_CSD_All_Grid.xlsm'):
        continue
    logName = logName.replace('_CSD_All_Grid.xlsm', '')
    MO_First_Row_Call1.append(logName)
    MO_First_Row_Call2.append(logName)
    MT_First_Row_Call1.append(logName)
    MT_First_Row_Call2.append(logName)
    cur_wb = load_workbook(CSD_file)
    mo_cur_ws = -1
    mo_rows = -1
    if 'VONR MO Call Setup Summary' in cur_wb.sheetnames:
        mo_cur_ws = cur_wb['VONR MO Call Setup Summary']
        mo_rows = mo_cur_ws.iter_rows(min_row=6, max_row=39, min_col=2, max_col=4)
    mt_cur_ws = -1
    mt_rows = -1
    if 'VONR MT Call Setup Summary' in cur_wb.sheetnames:
        mt_cur_ws = cur_wb['VONR MT Call Setup Summary']
        mt_rows = mt_cur_ws.iter_rows(min_row=6, max_row=35, min_col=2, max_col=4)
    
    if isFirstFile:
        if mo_rows != -1:
            for cols in mo_rows:
                MO_Summary_Sheet_Call1[cols[0].value] = [cols[0].value, '']
                MO_Summary_Sheet_Call1[cols[0].value].append(cols[1].value)
                MO_Summary_Sheet_Call2[cols[0].value] = [cols[0].value, '']
                MO_Summary_Sheet_Call2[cols[0].value].append(cols[2].value)
        if mt_rows != -1:
            for cols in mt_rows:
                MT_Summary_Sheet_Call1[cols[0].value] = [cols[0].value, '']
                MT_Summary_Sheet_Call1[cols[0].value].append(cols[1].value)
                MT_Summary_Sheet_Call2[cols[0].value] = [cols[0].value, '']
                MT_Summary_Sheet_Call2[cols[0].value].append(cols[2].value)
        isFirstFile = False
    else:
        if mo_rows != -1:
            for cols in mo_rows:
                MO_Summary_Sheet_Call1[cols[0].value].append(cols[1].value)
                MO_Summary_Sheet_Call2[cols[0].value].append(cols[2].value)
        if mt_rows != -1:
            for cols in mt_rows:
                MT_Summary_Sheet_Call1[cols[0].value].append(cols[1].value)
                MT_Summary_Sheet_Call2[cols[0].value].append(cols[2].value)       

def getAVG(row):
    totalCells = 0
    sum = 0
    for i in range(2, len(row)):
        if row[i] != 'N/A' and row[i] != None:
            sum += float(row[i])
            totalCells += 1
        else:
            continue
    if totalCells == 0:
        return sum
    else:
        return float(sum/totalCells)

wb_summary = Workbook()
ws_summary = wb_summary.active
ws_summary.title = 'MO_CSD_Summary_Call_1'
ws_summary.append(MO_First_Row_Call1)
for r_call1 in MO_Summary_Sheet_Call1.values():
    r_call1[1] = getAVG(r_call1)
    ws_summary.append(r_call1)

wb_summary.create_sheet('MO_CSD_Summary_Call_2')
ws_summary = wb_summary['MO_CSD_Summary_Call_2']
ws_summary.append(MO_First_Row_Call2)   
for r_call2 in MO_Summary_Sheet_Call2.values():
    r_call2[1] = getAVG(r_call2)
    ws_summary.append(r_call2)            

wb_summary.create_sheet('MT_CSD_Summary_Call_1')
ws_summary = wb_summary['MT_CSD_Summary_Call_1']
ws_summary.append(MT_First_Row_Call1)   
for r_call1 in MT_Summary_Sheet_Call1.values():
    r_call1[1] = getAVG(r_call1)
    ws_summary.append(r_call1)

wb_summary.create_sheet('MT_CSD_Summary_Call_2')
ws_summary = wb_summary['MT_CSD_Summary_Call_2']
ws_summary.append(MT_First_Row_Call2)   
for r_call2 in MT_Summary_Sheet_Call2.values():
    r_call2[1] = getAVG(r_call2)
    ws_summary.append(r_call2)  

dt_string = datetime.now().strftime('%Y%m%d_%H%M%S')
saveFileName = 'CSD_All_Logs_' + dt_string + '.xlsx'
savePath = os.path.join(ExtractCSD.workingDir, saveFileName)
print('CSD summary extracted: ' + savePath)
wb_summary.save(savePath)